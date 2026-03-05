
import pdfplumber
import re
import shutil
import os
import glob
from openpyxl import load_workbook
import difflib
import warnings
warnings.filterwarnings('ignore')

# Configuration
# Get the directory where this script is located
base_dir = os.path.dirname(os.path.abspath(__file__))
output_filename = "output.xlsx"
debug_file = os.path.join(base_dir, "debug_log_btg.txt")

def log_debug(msg):
    try:
        with open(debug_file, "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except:
        pass

def find_input_files():
    # Find PDF with BTG in name
    pdf_files = glob.glob(os.path.join(base_dir, "*BTG*.pdf"))
    if not pdf_files:
        raise FileNotFoundError("No BTG PDF files found in the directory.")
    pdf_files.sort(key=os.path.getmtime, reverse=True)
    pdf_path = pdf_files[0]
    
    # Check if output.xlsx exists - if so, use it as source (allows sequential processing)
    output_path = os.path.join(base_dir, "output.xlsx")
    if os.path.exists(output_path):
        print("Found existing output.xlsx - will update it in place")
        source_excel = output_path
    else:
        # Find Excel (excluding output files)
        excel_files = [f for f in glob.glob(os.path.join(base_dir, "*.xlsx")) 
                       if "output" not in os.path.basename(f).lower() 
                       and not os.path.basename(f).startswith("~$")]
        
        if not excel_files:
            raise FileNotFoundError("No input Excel files found.")
        
        excel_files.sort(key=os.path.getmtime, reverse=True)
        source_excel = excel_files[0]
    
    return pdf_path, source_excel

def clean_currency(value_str):
    if not value_str: return 0.0
    clean = value_str.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return 0.0

def clean_number(num_str):
    if not num_str: return 0.0
    clean = num_str.replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return None

def normalize_name(name):
    n = name.upper()
    n = n.replace("'", "").replace("'", "")
    n = n.replace("D'OR", "DOR").replace("D'OR", "DOR")
    
    n = re.sub(r'\b(PRECIFICACAO|PRECIFICAÇÃO|RENDA|A MERCADO|FIXA|08)\b', ' ', n)
    n = n.replace(":", " ")
    
    n = re.sub(r'\b(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[/\s]*20\d{2}\b', '', n)
    n = re.sub(r'IPC-?A\s*\+?\s*[\d,.]*%?', '', n)
    n = re.sub(r'[\d,.]+%\s*CDI', '', n)
    n = re.sub(r'CDI\s*\+?\s*[\d,.]*%?', '', n)
    n = re.sub(r'\d{2}/\d{2}/\d{2,4}', '', n)
    
    n = re.sub(r'\b(PRE|POS|FIM|FIC|FUNDO|INVESTIMENTO|MULTIMERCADO|RF|REND|CREDITO|PRIVADO|CP|LP|RL|RESP|LIMITADA|RESPONSABILIDADE|FIF)\b', ' ', n)
    n = n.replace('-', ' ').replace('_', ' ').replace('.', '').replace(',', '')
    
    final_n = " ".join(n.split())
    if final_n.startswith("DE "): final_n = final_n[3:]
    
    if "LUMINA" in name.upper():
        log_debug(f"Norm Debug: '{name}' -> '{final_n}'")
        
    return final_n

def extract_rate_pattern(text):
    """
    Extract rate pattern from text (e.g., "IPCA + 7,30%", "CDI + 2,5%", "101% CDI")
    Returns normalized rate string for matching
    """
    # Look for patterns like "IPCA + 7,30%" or "CDI + 2.5%"
    # Also handle split cases like "IPCA +" followed by "6.55"
    # Also handle CDI percentage like "101,00% do CDI" or "97% CDI"
    # Also handle Pre-fixed rates like "11,79% Pre" or "12.30%"
    patterns = [
        r'([\d,\.]+)\s*%\s*(?:do|da)?\s*(CDI|DI)',  # CDI percentage: 101% CDI or 101% do CDI
        r'(IPCA|CDI|DI|IGPM)\s*\+?\s*([\d,\.]+)\s*%',  # Standard format: IPCA + 7.30%
        r'(IPCA|CDI|DI|IGPM)\s*\+?\s*([\d,\.]+)',      # Without % symbol
        r'(IPCA|CDI|DI|IGPM)\s*[\+\s]+([\d,\.]+)',     # With + separator
        r'([\d,\.]+)\s*%\s*(?:Pre|Pré)',            # Explicit Pre-fixed: 11,79% Pre
        r'([\d,\.]+)\s*%\s*aa',                     # Explicit aa: 12,30% aa
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            # Check if it's a CDI percentage format (number comes first)
            if 'CDI' in pattern or 'DI' in pattern:
                if match.lastindex >= 2:
                    # It has groups for Rate and Index
                    p1, p2 = match.group(1), match.group(2)
                    # Check which one is the number
                    try:
                        rate_val = float(p1.replace(',', '.'))
                        index_type = p2.upper()
                        # CDI %
                        if 50 <= rate_val <= 150:
                            return f"{index_type}+{rate_val:.2f}"
                    except:
                        try:
                            rate_val = float(p2.replace(',', '.'))
                            index_type = p1.upper()
                            # Standard IPCA + X
                            if 0.01 <= rate_val <= 50:
                                return f"{index_type}+{rate_val:.2f}"
                        except:
                            pass
            elif 'Pre' in pattern or 'aa' in pattern:
                # Pre-fixed
                try:
                    rate_val = float(match.group(1).replace(',', '.'))
                    if 0.01 <= rate_val <= 30: # Reasonable pre-fixed range
                        return f"PRE+{rate_val:.2f}"
                except:
                    pass

    # Special case: look for "IPCA +" followed by a number nearby
    if re.search(r'(IPCA|CDI|DI|IGPM)\s*\+', text, re.IGNORECASE):
        # Find the index type
        index_match = re.search(r'(IPCA|CDI|DI|IGPM)', text, re.IGNORECASE)
        if index_match:
            index_type = index_match.group(1).upper()
            # Look for a percentage-like number after the index
            remaining_text = text[index_match.end():]
            number_match = re.search(r'[\+\s]+([\d,\.]+)', remaining_text)
            if number_match:
                rate_value = number_match.group(1).replace(',', '.')
                try:
                    rate_float = float(rate_value)
                    if 0.01 <= rate_float <= 150:
                        return f"{index_type}+{rate_float:.2f}"
                except:
                    pass
    
    # Fallback for just percentage if valid
    # Be careful not to match random percentages
    return None

def extract_pdf_data_btg(pdf_path):
    """
    Extract data from BTG PDF structure.
    BTG PDFs use a table-based format with different column arrangements.
    """
    extracted = []
    print(f"Extracting from: {os.path.basename(pdf_path)}")
    
    global_idx = 0
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Extract both tables and text
            tables = page.extract_tables()
            text = page.extract_text()
            
            if not text:
                continue
            
            # Process tables if available
            if tables:
                for table in tables:
                    for row in table:
                        if not row:
                            continue
                        
                        # Join row cells and process
                        row_text = ' '.join([str(cell) if cell else '' for cell in row])
                        
                        # Skip header/footer rows
                        if any(skip in row_text.lower() for skip in ['sac:', 'ouvidoria', 'total de', 'aplicações', 'resgates']):
                            continue
                        
                        # Look for investment data patterns
                        if process_btg_line(row_text, extracted, global_idx):
                            global_idx += 1
            
            # Also process raw text for lines not in tables
            # Use a smarter line buffer that detects investment boundaries
            lines = text.split('\n')
            line_buffer = ""
            
            for i, line in enumerate(lines):
                line_stripped = line.strip()
                if not line_stripped:
                    continue
                
                # Detect if this line starts a NEW investment entry
                # Look for: DEB, CRI, CRA, CDB, RDC, LCI, LCA, LFT, NTN, Funds, etc.
                is_new_investment = (
                    re.match(r'^[A-Z\s0-9]{15,}', line_stripped) or  # Long company/Fund name (15+ chars)
                    re.match(r'^(DEB|CRI|CRA|CDCA|LCI|LCA|CDB|RDC|LFT|NTN|NTNB|NTNF)', line_stripped, re.IGNORECASE) or  # Investment type
                    re.search(r'(DEB-|CRI-|CRA-|CDCA-|CDB-|RDC-|LFT|NTN)[A-Z0-9]*', line_stripped, re.IGNORECASE) or  # Investment code
                    'FUNDO' in line_stripped.upper() or 'FIC' in line_stripped.upper() or 'FIP' in line_stripped.upper() or  # Fund keywords
                    'HEADLINE' in line_stripped.upper() or 'CAPITAL' in line_stripped.upper() or 'EQUITY' in line_stripped.upper() or
                    'VENTURE' in line_stripped.upper() or 'ADVISORY' in line_stripped.upper() or 'MASTER' in line_stripped.upper() or
                    'SAFRA' in line_stripped.upper() or 'PICPAY' in line_stripped.upper() or 'BRB' in line_stripped.upper() or
                    'UNICRED' in line_stripped.upper() or 'SANTANDER' in line_stripped.upper()
                )
                
                if is_new_investment and line_buffer:
                    # Process the buffered investment before starting new one
                    if process_btg_line(line_buffer, extracted, global_idx):
                        global_idx += 1
                    line_buffer = line_stripped
                elif line_buffer:
                    # Check if this line is a continuation (has numbers/rates but no new investment marker)
                    has_data = (
                        re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', line_stripped) or
                        'IPCA' in line_stripped or 'CDI' in line_stripped or '%' in line_stripped or 'Pre' in line_stripped or 'Pré' in line_stripped
                    )
                    
                    if has_data:
                        # Append to current buffer
                        line_buffer += " " + line_stripped
                    else:
                        # Not data, might be noise - process buffer and start fresh
                        if process_btg_line(line_buffer, extracted, global_idx):
                            global_idx += 1
                        line_buffer = line_stripped
                else:
                    # Start new buffer
                    line_buffer = line_stripped
            
            # Process last buffered line
            if line_buffer:
                if process_btg_line(line_buffer, extracted, global_idx):
                    global_idx += 1
    
    return extracted

def process_btg_line(line, extracted, global_idx):
    """
    Process a single line from BTG PDF and extract investment data.
    Returns True if data was extracted, False otherwise.
    """
    # Skip empty or irrelevant lines
    if not line or not line.strip():
        return False
    
    lower_line = line.lower()
    if any(k in lower_line for k in ["relatório", "posição", "carteira", "sac:", "ouvidoria", "página", "total de", "aplicações", "resgates"]):
        return False
    
    # Look for lines with both dates and currency values
    dates = list(re.finditer(r'\d{2}/\d{2}/\d{2,4}', line))
    r_indices = [m.start() for m in re.finditer(r'R\$', line)]
    
    # Also look for lines with investment keywords and numbers
    # Expanded keywords to include Funds, CDBs, Government bonds
    keywords = ['IPCA', 'CDI', 'DEB', 'CRI', 'CRA', 'LCI', 'LCA', 'NTN', 'LFT', 'CDB', 'RDC', 'FUNDO', 'FIC', 'FIP', 'PRE', 'PRÉ', 
                'HEADLINE', 'CAPITAL', 'EQUITY', 'VENTURE', 'ADVISORY', 'MASTER', 'SAFRA', 'PICPAY', 'BRB', 'UNICRED', 'SANTANDER']
    has_investment_keyword = any(keyword in line.upper() for keyword in keywords)
    has_numbers = re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', line)
    
    # Relaxed condition: if we have keywords and numbers, or dates and numbers
    if (dates and has_numbers) or (r_indices and has_numbers) or (has_investment_keyword and has_numbers):
        # Extract name (usually at the beginning)
        name_part = ""
        
        if dates:
            first_date_idx = dates[0].start()
            name_part = line[:first_date_idx].strip()
        else:
            # Try to extract name before numbers
            # Use a more generous pattern for names (letters, spaces, special chars common in names)
            match = re.search(r'^([A-Z\s0-9\-\.\&]+(?:S\.?A\.?|LTDA|FIC|FIP|FUNDO)?)', line, re.IGNORECASE)
            if match:
                # Be careful not to capture the whole line if it starts with text
                # Stop at the first large number or keyword
                captured = match.group(1)
                # Find where numbers start
                num_start = re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', line)
                if num_start:
                    name_part = line[:num_start.start()].strip()
                else:
                    name_part = captured.strip()
        
        if not name_part or len(name_part) < 3:
            return False
        
        # Extract all monetary values - robustly
        # 1. Explicit R$ values
        all_money_strs = re.findall(r'R\$\s*[\d,.]+', line.replace('R$', ' R$'))
        all_values = [clean_currency(s) for s in all_money_strs]
        
        # 2. Extract numbers that look like currency (2 decimal places)
        # Only if we didn't find explicit R$ values
        if not all_values:
            potential_currency = re.findall(r'\b\d{1,3}(?:\.\d{3})*,\d{2}\b', line)
            all_values = [clean_number(s) for s in potential_currency]
            
        # Extract all numeric values (potential quantities)
        all_numbers = re.findall(r'\d{1,3}(?:\.\d{3})*,\d{2,8}', line)
        
        qty_candidates = []
        for num_str in all_numbers:
            val = clean_number(num_str)
            if val is not None and val > 0:
                # Filter out values that are clearly currency amounts
                if val < 1000000:  # Arbitrary threshold
                    qty_candidates.append(val)
        
        # Determine market value (Saldo Bruto R$) and purchase value (Valor Compra R$)
        # BTG PDFs: Saldo Bruto appears as the last/rightmost large value
        # Valor Compra R$ appears earlier in the sequence and matches Excel Quantity column
        # Pattern: [small_values..., valor_compra, ..., saldo_bruto, saldo_bruto]
        market_val = 0.0
        valor_compra = 0.0
        
        if all_values:
            # If we have R$ values, prefer them
            if len(all_values) >= 2:
                market_val = all_values[-2]  # Second to last is often the market value
                if len(all_values) >= 3:
                    valor_compra = all_values[-3]  # Third from last might be Valor Compra
            else:
                market_val = all_values[-1]
        elif qty_candidates:
            # No R$ values - extract from qty_candidates
            # Filter for large values (Saldo Bruto is typically > 1000)
            large_values = [v for v in qty_candidates if v > 1000]
            
            if large_values:
                # Saldo Bruto is typically the last large value
                # Often appears twice: [..., valor_compra, saldo_bruto, saldo_bruto]
                market_val = large_values[-1]
                
                # Valor Compra is often the second-to-last or third-to-last large value
                if len(large_values) >= 2:
                    # Check if last two values are the same (saldo_bruto appears twice)
                    if abs(large_values[-1] - large_values[-2]) < 0.01:
                        # They're the same, so valor_compra is before them
                        if len(large_values) >= 3:
                            valor_compra = large_values[-3]
                    else:
                        # Different values, second-to-last might be valor_compra
                        valor_compra = large_values[-2]
            elif qty_candidates:
                # Fallback: use largest value even if < 1000
                market_val = max(qty_candidates)
        
        # Extract rate pattern from the line for matching
        rate_pattern = extract_rate_pattern(line)
        
        # Only add if we have meaningful data
        if name_part and (market_val > 0 or qty_candidates):
            extracted.append({
                'id': global_idx,
                'name': name_part,
                'norm_name': normalize_name(name_part),
                'rate_pattern': rate_pattern,  # NEW: Rate pattern for matching
                'qty_candidates': qty_candidates,
                'value_candidates': all_values,
                'market_value': market_val,
                'valor_compra': valor_compra,  # NEW: Valor Compra R$
                'raw_line': line
            })
            
            log_debug(f"Extracted: {name_part} | MV: {market_val} | VC: {valor_compra} | Qty: {qty_candidates}")
            return True
        else:
             log_debug(f"Skipped (No Name/Val): {line[:50]}... | Name: {name_part} | MV: {market_val}")
             return False
    else:
        # debug helpful skip
        if any(k in line.upper() for k in keywords):
             log_debug(f"Skipped (No Criteria): {line[:50]}...")
    
    return False

def main():
    print("--- Starting BTG Auto-Reconciliation V1 ---")
    open(debug_file, 'w').close()
    
    try:
        pdf_path, source_excel = find_input_files()
        print(f"Input PDF: {os.path.basename(pdf_path)}")
        print(f"Input Excel: {os.path.basename(source_excel)}")
    except Exception as e:
        print(f"Error finding files: {e}")
        return

    target_excel = os.path.join(base_dir, output_filename)
    
    # If source and target are the same file, skip the copy step
    if os.path.abspath(source_excel) == os.path.abspath(target_excel):
        print(f"Updating {output_filename} in place...")
    else:
        # Source and target are different - copy the file
        try:
            if os.path.exists(target_excel):
                try:
                    os.write(target_excel, b'') 
                except:
                    pass 
                try:
                    os.rename(target_excel, target_excel)
                except OSError:
                    print(f"Warning: {output_filename} is locked. Saving to output_prox.xlsx")
                    target_excel = os.path.join(base_dir, "output_prox.xlsx")
                    
            shutil.copy2(source_excel, target_excel)
        except Exception as e:
            print(f"File copy failed: {e}")
            target_excel = os.path.join(base_dir, "output_forced.xlsx")
            shutil.copy2(source_excel, target_excel)

    pdf_rows = extract_pdf_data_btg(pdf_path)
    print(f"Extracted {len(pdf_rows)} rows.")
    
    print("Reading Excel...")
    wb = load_workbook(target_excel)
    ws = wb.active
    
    header_row = 12
    col_map = {'Name': 1, 'Qty': 2, 'Saldo': 3, 'SaldoExtrato': 4}
    for r in range(1, 40):
        row_values = []
        for c in ws[r]:
            val = str(c.value).strip() if c.value else ""
            row_values.append(val)
        row_str = " ".join(row_values)
        if "Quantidade" in row_str and "Saldo" in row_str:
            header_row = r
            for i, v in enumerate(row_values, 1):
                if 'Ativo' in v: col_map['Name'] = i
                elif 'Quantidade' in v: col_map['Qty'] = i
                elif v == 'Saldo': col_map['Saldo'] = i
                elif 'Saldo extrato' in v: col_map['SaldoExtrato'] = i
            break
            
    print(f"Header found at Row {header_row}. Mapping: {col_map}")
    
    excel_items = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell_qty = ws.cell(row=r, column=col_map['Qty'])
        cell_saldo = ws.cell(row=r, column=col_map['Saldo'])
        cell_name = ws.cell(row=r, column=col_map['Name'])
        cell_saldo_extrato = ws.cell(row=r, column=col_map['SaldoExtrato'])
        
        try:
             val_qty = cell_qty.value
             if val_qty is None: continue
             e_qty = float(val_qty)
             val_saldo = cell_saldo.value
             e_saldo = float(val_saldo) if val_saldo else 0.0
             e_name = str(cell_name.value).strip()
             
             # Skip if SaldoExtrato already has data
             existing_saldo_extrato = cell_saldo_extrato.value
             if existing_saldo_extrato is not None and existing_saldo_extrato != "" and existing_saldo_extrato != 0:
                 log_debug(f"Skipping row {r} - already has SaldoExtrato: {existing_saldo_extrato}")
                 continue
             
             # Extract rate pattern from Excel name
             e_rate_pattern = extract_rate_pattern(e_name)
             
             excel_items.append({
                 'row': r,
                 'name': e_name,
                 'rate_pattern': e_rate_pattern,  # NEW: Rate pattern from Excel

                 'norm_name': normalize_name(e_name),
                 'qty': e_qty,
                 'saldo': e_saldo,
                 'match': None,
                 'match_type': None
             })
        except:
             continue
             
    used_pdf_indices = set()
    updates = 0
    
    # Matching algorithm - RateMatch and ValorCompra are most reliable for BTG
    passes = ["RateMatch", "FundMatch", "StrictValue", "ValorCompra", "Strict", "ExactValue", "ExactName", "ApproxEVal", "FlexibleQty", "FinQty", "ApproxQty", "Value"]
    
    for pass_name in passes:
        for item in excel_items:
            if item['match']: continue 
            
            best_match = None
            
            if pass_name == "RateMatch":
                # Match based on rate pattern (IPCA + X.XX%, CDI + X.XX%)
                # This is very reliable for BTG PDFs as rates are unique identifiers
                if item['rate_pattern']:
                    for p in pdf_rows:
                        if p['id'] in used_pdf_indices: continue
                        if p.get('rate_pattern') and p['rate_pattern'] == item['rate_pattern']:
                            # Exact rate match - very reliable!
                            # Optional: check if values are in same ballpark
                            if p['market_value'] > 0 and item['saldo'] > 0:
                                ratio = p['market_value'] / item['saldo']
                                if 0.5 < ratio < 2.0:  # Within 2x range
                                    best_match = p
                                    break
                            else:
                                best_match = p
                                break
                                
            elif pass_name == "FundMatch":
                # Special pass for Funds where names vary slightly and values might differ by small % (e.g. net vs gross)
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    
                    # Calculate name similarity
                    score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                    
                    # STRICTER: Must have decent name similarity (> 0.5) AND close value match (< 3%)
                    if score > 0.5:
                        if p['market_value'] > 0 and item['saldo'] > 0:
                            diff_pct = abs(p['market_value'] - item['saldo']) / item['saldo']
                            if diff_pct < 0.03:  # Within 3%
                                best_match = p
                                break
                    
                    # For very strict value matching (< 0.1%), allow slightly lower similarity (> 0.4)
                    if p['market_value'] > 0 and item['saldo'] > 0:
                        diff_pct = abs(p['market_value'] - item['saldo']) / item['saldo']
                        if diff_pct < 0.001 and score > 0.4:
                             best_match = p
                             break

            elif pass_name == "StrictValue":
                # NEW: Very strict value matching (within 0.1%) with minimal name check
                # Helps catch duplicate investments with same rates
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    if p['market_value'] > 0 and item['saldo'] > 0:
                        diff_pct = abs(p['market_value'] - item['saldo']) / item['saldo']
                        if diff_pct < 0.001:  # Within 0.1%
                            # Very minimal name check - just needs some similarity
                            score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                            if score > 0.15:  # Very relaxed
                                best_match = p
                                break
                                
            elif pass_name == "ValorCompra":
                # Match based on Valor Compra R$ (matches Excel Quantity column)
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    if p['valor_compra'] > 0:
                        # Check if Valor Compra matches Excel Quantity (within small tolerance)
                        if abs(p['valor_compra'] - item['qty']) < 0.05:
                            # Optional: weak name similarity check to avoid wrong matches
                            score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                            if score > 0.2:  # Very relaxed name check
                                best_match = p
                                break
                            
            elif pass_name == "Strict":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    if any(abs(q - item['qty']) < 0.05 for q in p['qty_candidates']):
                         if p['norm_name'] == item['norm_name']:
                             best_match = p
                             break
                             
            elif pass_name == "ExactValue":
                for p in pdf_rows:
                     if p['id'] in used_pdf_indices: continue
                     if abs(p['market_value'] - item['saldo']) < 1.0: 
                         score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                         if score > 0.3:
                             best_match = p
                             break

            elif pass_name == "ExactName":
                 for p in pdf_rows:
                     if p['id'] in used_pdf_indices: continue
                     if p['norm_name'] == item['norm_name']:
                         best_match = p
                         break
                         
            elif pass_name == "ApproxEVal":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                    if score > 0.8:
                        diff = abs(p['market_value'] - item['saldo'])
                        if item['saldo'] > 0 and (diff / item['saldo']) < 0.03:
                            best_match = p
                            break

            elif pass_name == "FlexibleQty":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    for q in p['qty_candidates']:
                         if (abs(q - item['qty']*1000) < 0.05 or abs(q - item['qty']/1000) < 0.05 or
                             abs(q - item['qty']*100) < 0.05 or abs(q - item['qty']/100) < 0.05):
                            score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                            if score > 0.6: 
                                best_match = p
                                break
                    if best_match: break
                         
            elif pass_name == "FinQty":
                 if item['qty'] > 500:
                     for p in pdf_rows:
                         if p['id'] in used_pdf_indices: continue
                         if any(abs(v - item['qty']) < 5.0 for v in p['value_candidates']):
                             score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                             if score > 0.6:
                                 best_match = p
                                 break
                                 
            elif pass_name == "ApproxQty":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    for q in p['qty_candidates']:
                         if abs(q - item['qty']) < 1.0:
                             score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                             if score > 0.65:
                                 best_match = p
                                 break
                    if best_match: break
                    
            elif pass_name == "Value":
                 for p in pdf_rows:
                     if p['id'] in used_pdf_indices: continue
                     if abs(p['market_value'] - item['saldo']) < 5.0:
                         score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                         if score > 0.4:
                             best_match = p
                             break
            
            # --- VALIDATE ---
            if best_match:
                val = best_match['market_value']
                diff = abs(val - item['saldo'])
                valid = False
                
                if item['saldo'] == 0: 
                    valid = True
                elif (diff / item['saldo']) <= 0.20:
                    valid = True
                elif pass_name == "ExactValue": 
                    valid = True
                elif pass_name == "ApproxEVal":
                    valid = True
                elif pass_name == "Value" and diff < 10.0:
                    valid = True
                    
                # Exceptions
                if not valid:
                    if pass_name == "ExactName":
                         is_debt = any(k in item['norm_name'] for k in ["DEB", "CRI", "CRA", "LCI", "LCA", "NTN", "NB", "LETRAS", "TESOURO"])
                         if not is_debt: valid = True
                         
                if valid:
                    used_pdf_indices.add(best_match['id'])
                    item['match'] = val
                    item['match_type'] = pass_name
                    log_debug(f"MATCH: {item['name']} -> {best_match['name']} ({pass_name})")

    # --- WRITE ---
    for item in excel_items:
        if item['match'] is not None:
             cell_target = ws.cell(row=item['row'], column=col_map['SaldoExtrato'])
             cell_target.value = item['match']
             updates += 1
             
    wb.save(target_excel)
    print(f"Done. Updates: {updates}. Saved: {os.path.basename(target_excel)}")

if __name__ == "__main__":
    main()
