
import pdfplumber
import re
import shutil
import os
import glob
from openpyxl import load_workbook
import difflib

# Configuration
# Get the directory where this script is located
base_dir = os.path.dirname(os.path.abspath(__file__))
output_filename = "output.xlsx"
debug_file = os.path.join(base_dir, "debug_log.txt")

def log_debug(msg):
    try:
        with open(debug_file, "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except:
        pass

def find_input_files():
    # Find PDF with XP in name
    pdf_files = glob.glob(os.path.join(base_dir, "*XP*.pdf"))
    if not pdf_files:
        raise FileNotFoundError("No XP PDF files found in the directory.")
    pdf_files.sort(key=os.path.getmtime, reverse=True)
    pdf_path = pdf_files[0]
    
    # Check if output.xlsx exists - if so, use it as source (allows sequential processing)
    output_path = os.path.join(base_dir, "output.xlsx")
    if os.path.exists(output_path):
        print("Found existing output.xlsx - will update it in place")
        source_excel = output_path
    else:
        # Find Excel (excluding output files)
        excel_files = [f for f in glob.glob(os.path.join(base_dir, "*.xlsx")) 
                       if "output" not in os.path.basename(f).lower() 
                       and not os.path.basename(f).startswith("~$")]
        
        if not excel_files:
            raise FileNotFoundError("No input Excel files found.")
        
        excel_files.sort(key=os.path.getmtime, reverse=True)
        source_excel = excel_files[0]
    
    return pdf_path, source_excel

def clean_currency(value_str):
    if not value_str: return 0.0
    clean = value_str.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return 0.0

def clean_number(num_str):
    if not num_str: return 0.0
    clean = num_str.replace('.', '').replace(',', '.')
    try:
        return float(clean)
    except ValueError:
        return None

def normalize_name(name):
    n = name.upper()
    n = n.replace("’", "").replace("'", "")
    n = n.replace("D'OR", "DOR").replace("D’OR", "DOR")
    
    n = re.sub(r'\b(PRECIFICACAO|PRECIFICAÇÃO|RENDA|A MERCADO|FIXA|08)\b', ' ', n)
    n = n.replace(":", " ")
    
    n = re.sub(r'\b(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[/\s]*20\d{2}\b', '', n)
    n = re.sub(r'IPC-?A\s*\+?\s*[\d,.]*%?', '', n)
    n = re.sub(r'[\d,.]+%\s*CDI', '', n)
    n = re.sub(r'CDI\s*\+?\s*[\d,.]*%?', '', n)
    n = re.sub(r'\d{2}/\d{2}/\d{2,4}', '', n)
    
    n = re.sub(r'\b(PRE|POS|FIM|FIC|FUNDO|INVESTIMENTO|MULTIMERCADO|RF|REND|CREDITO|PRIVADO|CP|LP|RL|RESP|LIMITADA|RESPONSABILIDADE|FIF)\b', ' ', n)
    n = n.replace('-', ' ').replace('_', ' ').replace('.', '').replace(',', '')
    
    final_n = " ".join(n.split())
    if final_n.startswith("DE "): final_n = final_n[3:]
    
    if "LUMINA" in name.upper():
        log_debug(f"Norm Debug: '{name}' -> '{final_n}'")
        
    return final_n

def extract_pdf_data_universal(pdf_path):
    extracted = []
    print(f"Extracting from: {os.path.basename(pdf_path)}")
    
    with pdfplumber.open(pdf_path) as pdf:
        name_buffer = ""
        global_idx = 0
        
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if not text: continue
            
            lines = text.split('\n')
            for line in lines:
                dates = [m for m in re.finditer(r'\d{2}/\d{2}/\d{4}', line)]
                r_indices = [m.start() for m in re.finditer(r'R\$', line)]
                
                if dates and r_indices:
                    first_date_idx = dates[0].start()
                    name_part = line[:first_date_idx].strip()
                    full_name = (name_buffer + " " + name_part).strip()
                    name_buffer = ""
                    
                    last_date_idx = dates[-1].end()
                    first_money_idx = r_indices[0]
                    
                    qty_candidates = []
                    if first_money_idx > last_date_idx:
                        middle_segment = line[last_date_idx:first_money_idx].strip()
                        tokens = middle_segment.split()
                        for t in tokens:
                            if any(x in t for x in ['%', 'CDI', 'DI', 'IPCA', '+', '-', '/']):
                                continue
                            val = clean_number(t)
                            if val is not None:
                                qty_candidates.append(val)
                    
                    clean_line_money = line.replace('R$', ' R$')
                    all_money_strs = re.findall(r'R\$\s*[\d,.]+', clean_line_money)
                    all_values = [clean_currency(s) for s in all_money_strs]
                    
                    market_val = 0.0
                    if len(all_values) >= 2:
                        market_val = all_values[-2]
                    elif all_values:
                        market_val = all_values[-1]
                    
                    extracted.append({
                        'id': global_idx,
                        'name': full_name,
                        'norm_name': normalize_name(full_name),
                        'qty_candidates': qty_candidates,
                        'value_candidates': all_values,
                        'market_value': market_val,
                        'raw_line': line
                    })
                    global_idx += 1
                else:
                    lower_line = line.lower()
                    if any(k in lower_line for k in ["relatório", "posiç", "carteira", "ativo", "taxa a", "data cota", "valor cota", "mes", "download", "consulta", "página", "consolidada"]):
                        continue
                    name_buffer += " " + line.strip()
    return extracted

def main():
    print("--- Starting Auto-Reconciliation V20 (Factor-100 & ApproxValue) ---")
    open(debug_file, 'w').close()
    
    try:
        pdf_path, source_excel = find_input_files()
        print(f"Input PDF: {os.path.basename(pdf_path)}")
        print(f"Input Excel: {os.path.basename(source_excel)}")
    except Exception as e:
        print(f"Error finding files: {e}")
        return

    target_excel = os.path.join(base_dir, output_filename)
    
    # If source and target are the same file, skip the copy step
    if os.path.abspath(source_excel) == os.path.abspath(target_excel):
        print(f"Updating {output_filename} in place...")
    else:
        # Source and target are different - copy the file
        try:
            if os.path.exists(target_excel):
                try:
                    os.write(target_excel, b'') 
                except:
                    pass 
                try:
                    os.rename(target_excel, target_excel)
                except OSError:
                    print(f"Warning: {output_filename} is locked. Saving to output_prox.xlsx")
                    target_excel = os.path.join(base_dir, "output_prox.xlsx")
                    
            shutil.copy2(source_excel, target_excel)
        except Exception as e:
            print(f"File copy failed: {e}")
            target_excel = os.path.join(base_dir, "output_forced.xlsx")
            shutil.copy2(source_excel, target_excel)

    pdf_rows = extract_pdf_data_universal(pdf_path)
    print(f"Extracted {len(pdf_rows)} rows.")
    
    print("Reading Excel...")
    wb = load_workbook(target_excel)
    ws = wb.active
    
    header_row = 12
    col_map = {'Name': 1, 'Qty': 2, 'Saldo': 3, 'SaldoExtrato': 4}
    for r in range(1, 40):
        row_values = []
        for c in ws[r]:
            val = str(c.value).strip() if c.value else ""
            row_values.append(val)
        row_str = " ".join(row_values)
        if "Quantidade" in row_str and "Saldo" in row_str:
            header_row = r
            for i, v in enumerate(row_values, 1):
                if 'Ativo' in v: col_map['Name'] = i
                elif 'Quantidade' in v: col_map['Qty'] = i
                elif v == 'Saldo': col_map['Saldo'] = i
                elif 'Saldo extrato' in v: col_map['SaldoExtrato'] = i
            break
            
    print(f"Header found at Row {header_row}. Mapping: {col_map}")
    
    excel_items = []
    for r in range(header_row + 1, ws.max_row + 1):
        cell_qty = ws.cell(row=r, column=col_map['Qty'])
        cell_saldo = ws.cell(row=r, column=col_map['Saldo'])
        cell_name = ws.cell(row=r, column=col_map['Name'])
        
        try:
             val_qty = cell_qty.value
             if val_qty is None: continue
             e_qty = float(val_qty)
             val_saldo = cell_saldo.value
             e_saldo = float(val_saldo) if val_saldo else 0.0
             e_name = str(cell_name.value).strip()
             
             excel_items.append({
                 'row': r,
                 'name': e_name,
                 'norm_name': normalize_name(e_name),
                 'qty': e_qty,
                 'saldo': e_saldo,
                 'match': None,
                 'match_type': None
             })
        except:
             continue
             
    used_pdf_indices = set()
    updates = 0
    
    passes = ["Strict", "ExactValue", "ExactName", "ApproxEVal", "FlexibleQty", "FinQty", "ApproxQty", "Value"]
    
    for pass_name in passes:
        for item in excel_items:
            if item['match']: continue 
            
            best_match = None
            
            if pass_name == "Strict":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    if any(abs(q - item['qty']) < 0.05 for q in p['qty_candidates']):
                         if p['norm_name'] == item['norm_name']:
                             best_match = p
                             break
                             
            elif pass_name == "ExactValue":
                for p in pdf_rows:
                     if p['id'] in used_pdf_indices: continue
                     if abs(p['market_value'] - item['saldo']) < 1.0: 
                         score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                         if score > 0.3:
                             best_match = p
                             break

            elif pass_name == "ExactName":
                 for p in pdf_rows:
                     if p['id'] in used_pdf_indices: continue
                     if p['norm_name'] == item['norm_name']:
                         best_match = p
                         break
                         
            elif pass_name == "ApproxEVal":
                # APPROX VALUE MATCH (New for V20)
                # Finds Items where NormNames are identical/very close 
                # AND Values are within 3% (e.g. 151k vs 152k)
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    # Check Name similarity first (Efficient)
                    score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                    if score > 0.8: # Must be almost same name
                        # Check Value Divergence
                        diff = abs(p['market_value'] - item['saldo'])
                        if item['saldo'] > 0 and (diff / item['saldo']) < 0.03: # 3% only
                            best_match = p
                            break

            elif pass_name == "FlexibleQty":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    for q in p['qty_candidates']:
                         # Factor 1000 or 100
                         if (abs(q - item['qty']*1000) < 0.05 or abs(q - item['qty']/1000) < 0.05 or
                             abs(q - item['qty']*100) < 0.05 or abs(q - item['qty']/100) < 0.05):
                            score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                            if score > 0.6: 
                                best_match = p
                                break
                    if best_match: break
                         
            elif pass_name == "FinQty":
                 if item['qty'] > 500:
                     for p in pdf_rows:
                         if p['id'] in used_pdf_indices: continue
                         if any(abs(v - item['qty']) < 5.0 for v in p['value_candidates']):
                             score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                             if score > 0.6:
                                 best_match = p
                                 break
                                 
            elif pass_name == "ApproxQty":
                for p in pdf_rows:
                    if p['id'] in used_pdf_indices: continue
                    for q in p['qty_candidates']:
                         if abs(q - item['qty']) < 1.0:
                             score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                             if score > 0.65:
                                 best_match = p
                                 break
                    if best_match: break
                    
            elif pass_name == "Value":
                 for p in pdf_rows:
                     if p['id'] in used_pdf_indices: continue
                     if abs(p['market_value'] - item['saldo']) < 5.0:
                         score = difflib.SequenceMatcher(None, item['norm_name'], p['norm_name']).ratio()
                         if score > 0.4:
                             best_match = p
                             break
            
            # --- VALIDATE ---
            if best_match:
                val = best_match['market_value']
                diff = abs(val - item['saldo'])
                valid = False
                
                if item['saldo'] == 0: 
                    valid = True
                elif (diff / item['saldo']) <= 0.20:
                    valid = True
                elif pass_name == "ExactValue": 
                    valid = True
                elif pass_name == "ApproxEVal": # Already validated 3% inside loop
                    valid = True
                elif pass_name == "Value" and diff < 10.0:
                    valid = True
                    
                # Exceptions
                if not valid:
                    if pass_name == "ExactName":
                         is_debt = any(k in item['norm_name'] for k in ["DEB", "CRI", "CRA", "LCI", "LCA", "NTN", "NB", "LETRAS", "TESOURO"])
                         if not is_debt: valid = True
                         
                if valid:
                    used_pdf_indices.add(best_match['id'])
                    item['match'] = val
                    item['match_type'] = pass_name
                    if "LUMINA" in item['name'].upper() or "ENERGISA" in item['name'].upper() or "AZ QUEST" in item['name'].upper() or "IRANI" in item['name'].upper():
                        log_debug(f"MATCH: {item['name']} -> {best_match['name']} ({pass_name})")

    # --- WRITE ---
    for item in excel_items:
        if item['match'] is not None:
             cell_target = ws.cell(row=item['row'], column=col_map['SaldoExtrato'])
             cell_target.value = item['match']
             updates += 1
             
    wb.save(target_excel)
    print(f"Done. Updates: {updates}. Saved: {os.path.basename(target_excel)}")

if __name__ == "__main__":
    main()
